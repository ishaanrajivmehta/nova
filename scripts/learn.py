#!/usr/bin/env python3
"""
learn.py — the RECURSIVE layer (Sir, 8 Jul 2026): as more JDs land, the data + evidence get richer,
so raw scores climb and each craft needs less lift over time.

Each run mines the whole JD corpus in the backend and produces:
  1. knowledge/corpus-insights.md — rolling "what the market wants" report (top skills by JD frequency),
     the SKILL GAPS (skills the market keeps asking for that Sir's evidence bank doesn't cover = honest
     learn-targets), and the stretch skills most often demanded.
  2. engine/data/gazetteer_candidates.json — new skill-like terms seen in JDs that aren't in the
     gazetteer yet, for review. HIGH-CONFIDENCE candidates (frequent + tech-shaped) are AUTO-PROMOTED into
     the live gazetteer; the rest are parked for the tailoring task (Claude) to promote with judgement.

Safety: this only ever grows the KEYWORD DATA (gazetteer/synonyms) and writes reports. It NEVER writes
new EVIDENCE bullets — those stay truth-reviewed (the tailoring task folds in well-scoring reframings by
hand). So the recursion cannot fabricate a checkable fact.

CLI:  python3 engine/learn.py            # mine + report + auto-promote high-confidence terms
      python3 engine/learn.py --dry      # report only, no writes to the gazetteer
"""
from __future__ import annotations
import json, os, re, sys, argparse
from collections import Counter
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import cv_score

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
_DATA = os.path.join(_HERE, "data")
_BANK = os.path.join(_HERE, "evidence_bank.json")
_BACKEND = os.path.join(_ROOT, "Job-Hunt-Backend.xlsx")
_REPORT = os.path.join(_ROOT, "knowledge", "corpus-insights.md")
_CANDS = os.path.join(_DATA, "gazetteer_candidates.json")

PROMOTE_PCT = 12          # min % of JDs a term must appear in before auto-promotion (was 8)

# a tech-shaped filter so we don't auto-add generic English as "skills"
# 27 Jul 2026 — FIXED: these were substring matches, so "bi" matched aBIlity / responsiBIlities /
# reliaBIlity and "analy" matched "analyst london". 38 n-gram fragments reached hard_skills at
# BASE_WEIGHT 1.0 and distorted JD-match by up to 22 points. Short hints are now token-anchored.
_TECH_STEM = re.compile(r"(sql|python|java|api|aws|azure|gcp|cloud|data|model|pipeline|analy|dashboard|"
                        r"etl|spark|kafka|docker|kubernet|react|node|\.js|power|tableau|looker|snowflake|"
                        r"bigquery|airflow|dbt|pytorch|tensor|nlp|llm|rag|scikit|pandas|numpy|git|ci/cd|"
                        r"stakeholder|forecast|regression|cluster|statistic|visuali|warehouse|databricks|"
                        # 27 Jul 2026 — broadened beyond the original data/AI-only stem list, which was
                        # parking legitimate tools (terraform, salesforce, hubspot) as "not tech-shaped".
                        r"terraform|ansible|jenkins|linux|bash|devops|infra|serverless|lambda|redis|"
                        r"postgres|mysql|mongo|elastic|graphql|rest|micro|oauth|saas|crm|erp|"
                        r"salesforce|hubspot|workday|sap|oracle|dynamics|servicenow|zendesk|jira|"
                        r"figma|adobe|excel|sheets|notion|slack|automat|integrat|migrat|deploy|"
                        r"test|agile|scrum|governance|compliance|security|encrypt|audit)", re.I)
# short hints must be a WHOLE token — never a substring of an ordinary English word
_TECH_TOKEN = {"ml", "ai", "bi", "r", "go", "c", "c#", "c++", "js", "ts", "qa", "ux", "ui"}

def _tech_shaped(term: str) -> bool:
    toks = term.lower().split()
    return bool(_TECH_STEM.search(term)) or any(t in _TECH_TOKEN for t in toks)

# ---- PROMOTION GATE (27 Jul 2026) -----------------------------------------------------------
# Frequency alone is NOT evidence that a phrase is a skill. Anything failing this gate is parked
# in gazetteer_candidates.json for human/Claude review instead of entering the live gazetteer.
_FUNCTION = {"the","a","an","of","in","with","and","to","for","from","on","at","by","or","as","is",
             "are","be","ability","strong","key","other","more","most","including","etc","your",
             "you","our","we","their","this","that","will","have","has","who","what","across"}
_ROLE = {"analyst","engineer","developer","manager","scientist","consultant","specialist","associate",
         "director","lead","architect","administrator","officer","coordinator","executive","intern",
         "advisor","assistant","head","partner","owner"}
_GENERIC = {"data","analysis","analytical","analyse","analyze","model","models","modelling","modeling",
            "cloud","datasets","dataset","responsibilities","responsibility","power","maintain",
            "training","ability","statistical","stakeholder","stakeholders","databases","database",
            "reliability","experience","skills","knowledge","team","business","technology","tools",
            "systems","projects","support","quality","process","solutions","services","environment"}

def curated_protected(gaz: dict, syn: dict) -> set:
    """Terms exempt from the gate: canonical synonym keys plus the gazetteer's explicit
    `protected_terms`. Curated entries were reviewed by a human — the gate exists to stop
    UNKNOWN mined n-grams, and applying it to reviewed material strips real skills
    ('machine learning', 'lead scoring', '5s', 'theory of constraints')."""
    out = {str(k).lower() for k in syn if not str(k).startswith("_")}
    out |= {str(t).lower() for t in gaz.get("protected_terms", [])}
    return out

def structural_reject_reason(term: str, protected: set) -> str | None:
    """Is this string even shaped like a skill? Used to AUDIT the existing curated gazetteer.
    Deliberately does NOT require tech-shape — curated entries were chosen by a human, and plenty
    of real skills ('machine learning', 'user stories') match no tech stem."""
    s = term.strip().lower()
    if s in protected:                       return None      # canonical / explicitly protected
    if s != term:                            return "whitespace"
    if len(s) < 3 and s not in _TECH_TOKEN:  return "too short"
    if not re.match(r"^[a-z0-9][a-z0-9 +/#.&-]*$", s):        return "illegal characters"
    if s.rstrip(".,;:!?") != s:              return "trailing punctuation"
    toks = s.split()
    if len(toks) > 4:                        return "too long (>4 words)"
    if any(w in _FUNCTION for w in toks):    return "contains function word"
    if s in _ROLE or any(t in _ROLE for t in toks):  return "role word (never a skill)"
    if s in _GENERIC:                        return "bare generic noun"
    return None

def promotion_reject_reason(term: str, protected: set) -> str | None:
    """Stricter gate for AUTO-PROMOTING an unknown term mined from the corpus. Structural checks
    PLUS a tech-shape requirement: an unrecognised phrase must look like a tool/technology before it
    is allowed to carry hard-skill weight. Anything rejected is parked for review, never discarded."""
    why = structural_reject_reason(term, protected)
    if why:                                  return why
    if term.strip().lower() in protected:    return None
    if not _tech_shaped(term):               return "not tech-shaped"
    return None

def _load_jds():
    import openpyxl
    wb = openpyxl.load_workbook(_BACKEND, data_only=True); ws = wb["Jobs"]
    hdr = [c.value for c in ws[1]]; ix = {h: i for i, h in enumerate(hdr)}
    jds = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        jd = r[ix["JD Text"]]
        if jd and len(str(jd)) > 250:
            jds.append(str(jd))
    return jds

def corpus_stats(jds):
    syn, _, gaz = cv_score._data()
    allskills = gaz["hard_skills"] + gaz["soft_skills"]
    n = max(len(jds), 1)
    cnt = Counter()
    for jd in jds:
        jn = cv_score.norm(jd); seen = set()
        for s in allskills:
            surfaces = [s] + syn.get(cv_score.canon(s), [])
            if any(cv_score._contains(jn, x) for x in surfaces):
                seen.add(cv_score.canon(s))
        for s in seen:
            cnt[s] += 1
    return [(s, c, round(100 * c / n)) for s, c in cnt.most_common()], n

def sir_skills():
    bank = json.load(open(_BANK, encoding="utf-8"))
    pool = set()
    for role in bank["roles"]:
        for b in role["bullets"]:
            pool |= {cv_score.canon(s) for s in b.get("skills", [])}
    for cat, items in bank["skills_bank"].items():
        pool |= {cv_score.canon(s) for s in items}
    pool |= {cv_score.canon(s) for s in bank["stretch_skills"]["items"]}
    return pool

def candidate_terms(jds, top=40):
    """frequent tech-shaped bigrams/tokens not already a known skill/synonym."""
    syn, _, gaz = cv_score._data()
    known = set()
    for s in gaz["hard_skills"] + gaz["soft_skills"] + gaz["qualifications"]:
        known.add(cv_score.canon(s))
    for c, vs in syn.items():
        known.add(c); known |= {v.lower() for v in vs}
    n = max(len(jds), 1); cnt = Counter()
    for jd in jds:
        toks = cv_score.norm(jd).split()
        grams = set(toks) | {f"{toks[i]} {toks[i+1]}" for i in range(len(toks) - 1)}
        for g in grams:
            if g in known or len(g) < 3 or g.isdigit():
                continue
            if _tech_shaped(g):
                cnt[g] += 1
    out = [(g, c, round(100 * c / n)) for g, c in cnt.most_common() if c >= max(3, n // 20)]
    return out[:top]

def run(dry=False):
    jds = _load_jds()
    stats, n = corpus_stats(jds)
    have = sir_skills()
    gaps = [(s, pct) for s, c, pct in stats if cv_score.canon(s) not in have]
    cands = candidate_terms(jds)

    # auto-promote high-confidence candidates into the live gazetteer.
    # 27 Jul 2026: threshold raised 8% -> 12% AND every candidate must pass promotion_reject_reason().
    # Rejects are PARKED for review, never silently promoted. Frequency is not evidence of skill-hood.
    promoted, rejected = [], []
    if not dry and cands:
        gaz = cv_score._load_json("skills_gazetteer.json")
        syn, _, _ = cv_score._data()
        protected = curated_protected(gaz, syn)
        existing = set(x.lower() for x in gaz["hard_skills"])
        for term, c, pct in cands:
            if term in existing:
                continue
            why = promotion_reject_reason(term, protected)
            if why:
                rejected.append({"term": term, "jd_pct": pct, "rejected": why})
            elif pct >= PROMOTE_PCT:
                gaz["hard_skills"].append(term); promoted.append((term, pct)); existing.add(term)
            else:
                rejected.append({"term": term, "jd_pct": pct, "rejected": f"below {PROMOTE_PCT}% threshold"})
        if promoted:
            json.dump(gaz, open(os.path.join(_DATA, "skills_gazetteer.json"), "w"), indent=2)
        json.dump({"generated": True,
                   "promote_threshold_pct": PROMOTE_PCT,
                   "auto_promoted": [{"term": t, "jd_pct": p} for t, p in promoted],
                   "parked_for_review": rejected,
                   "candidates": [{"term": t, "jd_pct": p} for t, c, p in cands]},
                  open(_CANDS, "w"), indent=2)

    # write the rolling report
    L = [f"# Corpus insights — {n} JDs (auto-generated by engine/learn.py)", "",
         "Rolling view of what Sir's target market asks for. Feeds master-CV emphasis + the tailoring loop.",
         "As the corpus grows this sharpens and the raw-score baseline rises (less lift per craft).", "",
         "## Top skills demanded (% of JDs)"]
    for s, c, pct in stats[:40]:
        L.append(f"- {pct:3d}%  {s}")
    L += ["", "## Skill gaps — market wants, evidence bank lacks (honest learn-targets)"]
    L += [f"- {pct:3d}%  {s}" for s, pct in gaps[:20]] or ["- (none — bank covers the top demand)"]
    L += ["", "## New candidate terms seen in JDs (for gazetteer review)"]
    L += [f"- {p:3d}%  {t}" for t, c, p in cands[:25]] or ["- (none)"]
    if promoted:
        L += ["", "## Auto-promoted into the gazetteer this run", *[f"- {t} ({p}%)" for t, p in promoted]]
    os.makedirs(os.path.dirname(_REPORT), exist_ok=True)
    open(_REPORT, "w", encoding="utf-8").write("\n".join(L) + "\n")
    print(f"[learn] {n} JDs · top skill: {stats[0][0] if stats else '—'} · "
          f"{len(gaps)} gaps · {len(cands)} candidates · {len(promoted)} auto-promoted → {_REPORT}")

if __name__ == "__main__":
    ap = argparse.ArgumentParser(); ap.add_argument("--dry", action="store_true"); a = ap.parse_args()
    run(dry=a.dry)
